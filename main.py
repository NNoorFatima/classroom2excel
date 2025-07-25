import pandas as pd
import json
import re
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

# --- Step 1: Load gcr_marks.txt as JSON list ---
with open("gcr_marks.txt", "r", encoding="utf-8") as f:
    gcr_data = json.load(f)

# Build lookup dictionaries
marks_by_last4 = {}
marks_by_name = {}
for entry in gcr_data:
    last4 = entry.get("roll_number_last4")
    name = entry.get("name", "").strip().lower()
    mark = entry.get("marks", "0/100").split("/")[0]  # extract numeric part

    if last4:
        marks_by_last4[last4] = mark
    if name:
        marks_by_name[name] = mark

# --- Step 2: Load Excel sheet ---
df = pd.read_excel("section A.xlsx")  # Adjust filename if needed

# Normalize column names
df.columns = [col.strip() for col in df.columns]
if "RollNo" not in df.columns or "Name" not in df.columns:
    raise Exception("Excel must have columns: RollNo, Name")

# --- Step 3: Add marks column ---
marks = []
no_match_rows = []

for idx, row in df.iterrows():
    roll = str(row["RollNo"]).strip()
    name = str(row["Name"]).strip().lower()

    # Extract last 4 digits of roll
    match = re.search(r"(\d{4})\b", roll)
    last4 = match.group(1) if match else None

    mark = None
    if last4 and last4 in marks_by_last4:
        mark = marks_by_last4[last4]
    elif name in marks_by_name:
        mark = marks_by_name[name]

    if mark:
        marks.append(int(mark))
    else:
        marks.append("")  # leave blank
        no_match_rows.append(idx)

df["Marks"] = marks

# --- Step 4: Save + Highlight no match rows red ---
output_file = "output_with_marks.xlsx"
df.to_excel(output_file, index=False)

# Highlight red
wb = load_workbook(output_file)
ws = wb.active
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
for row in no_match_rows:
    for col in range(1, ws.max_column + 1):
        ws.cell(row=row + 2, column=col).fill = red_fill  # +2 for header and 1-index

wb.save(output_file)
print("✅ Excel updated with marks and saved to output_with_marks.xlsx")
